from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from accounts.models import Customer
from transactions.models import Deposit, Withdrawal
from django.db.models import Sum
from django.contrib.auth.decorators import login_required
from .forms import DepositForm, WithdrawalForm
from .email_system import creditMessage, debitMessage
from .filters import DepositTransactionFilter, WithdrawalTransactionFilter
import xlwt
from datetime import datetime
from django.contrib import messages
from openpyxl import Workbook, utils

# Create your views here.
def home(request):
    if not request.user.is_authenticated:
        return render(request, "home.html")
    else:
        customer = Customer.objects.get(user=request.user)
        deposit = Deposit.objects.filter(customer=customer)
        deposit_total = deposit.aggregate(Sum('amount'))['amount__sum']
        withdrawal = Withdrawal.objects.filter(customer=customer)
        withdrawal_total = withdrawal.aggregate(Sum('amount'))['amount__sum']
        context = {
            'customer':customer,
            'deposit': deposit,
            'deposit_total':deposit_total,
            'withdrawal': withdrawal,
            'withdrawal_total': withdrawal_total
        }

        return render(request, 'transactions.html', context)

def about(request):
    return render(request, 'about.html')

@login_required
def depositView(request):
    context = {
        'title':'Deposit'
    }
    messages=[]
    if request.method == 'POST':
        form = DepositForm(request.POST)
        if form.is_valid():
            if int(request.POST['amount']) <= 100000 and int(request.POST['amount']) > 0:
                deposit = form.save(commit=False)
                customer = get_object_or_404(Customer, user=request.user)
                deposit.customer = customer
                deposit.save()

                customer.balance += deposit.amount
                customer.save()
                # creditMessage(customer.account_no, deposit.amount, customer.balance, customer.email)
                return redirect('/')

            else:
                messages = ["Please enter a valid amount between 1 to 10000000"]
        context['messages'] = messages
        context['form'] = form
        context['balance'] = get_object_or_404(Customer, user=request.user).balance
        return render(request, 'transactions/form.html', context)
        
    else:
        form = DepositForm()

        context['form'] = form
        context['balance'] = get_object_or_404(Customer, user=request.user).balance
        return render(request, "transactions/form.html", context)


@login_required
def withdrawalView(request):
    context = {
        'title':'Withdraw'
    }
    # messages=[]
    if request.method == 'POST':
        form = WithdrawalForm(request.POST)
        print(form.errors)
        if form.is_valid():
            print(int(request.POST['amount']))
            if int(request.POST['amount']) <= 100000 and int(request.POST['amount']) > 0:
                withdraw = form.save(commit=False)
                customer = get_object_or_404(Customer, user=request.user)
                if withdraw.amount <= customer.balance:
                    
                    withdraw.customer = customer
                    withdraw.save()
                    customer.balance -= withdraw.amount
                    customer.save()
                    # debitMessage(customer.account_no, withdraw.amount, customer.balance, customer.email)
                    return redirect('/')
                else:
                    messages.add_message(request, messages.ERROR, "You can not withdraw amount more than the balance in your account.")

            else:
                messages.add_message(request, messages.WARNING, "Please enter a valid amount between 1 to 10000000.")
            # context['messages'] = messages
        context['form'] = form
        context['balance'] = get_object_or_404(Customer, user=request.user).balance
        return render(request, 'transactions/form.html', context)
    else:
        form = WithdrawalForm()

        context['form'] = form
        context['balance'] = get_object_or_404(Customer, user=request.user).balance
        return render(request, "transactions/form.html", context)

@login_required
def recordSheetView(request):

    customer = Customer.objects.get(user=request.user)
    deposit = Deposit.objects.filter(customer=customer)
    withdrawal = Withdrawal.objects.filter(customer=customer)
    depositTransactionFilter = DepositTransactionFilter(request.GET, queryset=deposit)
    withdrawalTransactionFilter = WithdrawalTransactionFilter(request.GET, queryset=withdrawal)
    deposit = depositTransactionFilter.qs
    print(depositTransactionFilter)
    
    # response = export_excel(deposit)

    context = {
        'withdrawalTransactionFilter': withdrawalTransactionFilter,
        'depositTransactionFilter': depositTransactionFilter,
        'deposit': deposit,
        'withdrawal': withdrawal,
    }
    return render(request, 'recordSheet.html', context)

@login_required
def export_excel(request, type):
    print(request.GET)
    qs=None
    if type=='Deposit':
        customer = Customer.objects.get(user=request.user)
        qs = Deposit.objects.filter(customer=customer)
    elif type=='Withdrawal':
        customer = Customer.objects.get(user=request.user)
        qs = Withdrawal.objects.filter(customer=customer)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    response['Content-Disposition'] = 'attachment; filename={date}-DebitTransaction.xlsx'.format(date=datetime.now().strftime('%Y-%m-%d'),)
    workbook = Workbook()

    print()
    worksheet = workbook.active
    accountDetails = [['Name', customer.name], ['Account No.', customer.account_no]]
    
    worksheet.cell(1,1).value = 'Name'
    worksheet.cell(1,2).value = customer.name
    worksheet.cell(2,1).value = 'Account No.'
    worksheet.cell(2,2).value = customer.account_no
    worksheet.cell(3,1).value = 'Current Balance'
    worksheet.cell(3,2).value = 'Rs. ' + str(customer.balance)


    columns = [
        
        'Date',
        'Type',
        'Amount (Rs.)',
    ]

    row_num = 4

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        worksheet.column_dimensions[utils.get_column_letter(col_num)].width = 30
    for q in qs:
        row_num += 1
        
        # Define the data for each cell in the row 
        row = [
            q.timestamp.strftime('%Y-%m-%d'),
            type,
            q.amount,
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
    
    workbook.save(response)
    return response
